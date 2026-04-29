#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Callable, Iterable

import pandas as pd
from dateutil import parser as date_parser
from pypdf import PdfReader


DATE_HINTS = ("giorno", "mese", "anno")
ITALIAN_MONTHS = {
    "gennaio": "January",
    "febbraio": "February",
    "marzo": "March",
    "aprile": "April",
    "maggio": "May",
    "giugno": "June",
    "luglio": "July",
    "agosto": "August",
    "settembre": "September",
    "ottobre": "October",
    "novembre": "November",
    "dicembre": "December",
}
MONEY_LABELS = {
    "total_amount_eur": [
        "totale da pagare",
        "totale bolletta",
        "importo da pagare",
        "totale fattura",
        "totale documento",
    ],
    "taxes_eur": ["imposte", "accise"],
    "vat_eur": ["iva"],
    "transport_eur": ["spesa per il trasporto", "trasporto e gestione del contatore"],
    "system_charges_eur": ["spesa per oneri di sistema", "oneri di sistema"],
    "energy_cost_eur": ["spesa per la materia energia", "costo energia", "corrispettivo energia"],
    "tv_license_eur": ["canone rai", "canone tv"],
}
TEXT_LABELS = {
    "supplier_name": ["venditore", "fornitore", "societa di vendita", "emittente"],
    "invoice_number": ["numero fattura", "n. fattura", "fattura n", "numero documento"],
    "invoice_date": ["data emissione", "data fattura", "emessa il"],
    "due_date": ["scadenza", "data scadenza"],
    "customer_name": ["intestatario", "cliente", "ragione sociale", "nominativo"],
    "supply_address": ["indirizzo di fornitura", "presso fornitura", "ubicazione fornitura"],
    "tariff_code": ["offerta", "tipologia offerta", "tariffa", "codice offerta"],
    "pod_code": ["pod"],
}
PERIOD_LABELS = [
    "periodo",
    "periodo di riferimento",
    "consumi dal",
    "periodo fatturato",
]

OUTPUT_COLUMNS = [
    "source_file",
    "supplier_template",
    "supplier_name",
    "invoice_number",
    "invoice_date",
    "due_date",
    "customer_name",
    "supply_address",
    "pod_code",
    "tariff_code",
    "billing_period_start",
    "billing_period_end",
    "consumption_kwh",
    "consumption_f1_kwh",
    "consumption_f2_kwh",
    "consumption_f3_kwh",
    "committed_power_kw",
    "total_amount_eur",
    "invoice_total_eur",
    "bonus_eur",
    "energy_cost_eur",
    "transport_eur",
    "system_charges_eur",
    "taxes_eur",
    "vat_eur",
    "tv_license_eur",
    "energy_qty",
    "energy_unit_rate",
    "energy_imponibile_eur",
    "losses_qty",
    "losses_unit_rate",
    "losses_imponibile_eur",
    "dispbt_qty",
    "dispbt_unit_rate",
    "dispbt_imponibile_eur",
    "commercialization_qty",
    "commercialization_unit_rate",
    "commercialization_imponibile_eur",
    "capacity_market_qty",
    "capacity_market_unit_rate",
    "capacity_market_imponibile_eur",
    "dispatching_qty",
    "dispatching_unit_rate",
    "dispatching_imponibile_eur",
    "transport_energy_qty",
    "transport_energy_unit_rate",
    "transport_energy_imponibile_eur",
    "transport_fixed_qty",
    "transport_fixed_unit_rate",
    "transport_fixed_imponibile_eur",
    "transport_power_qty",
    "transport_power_unit_rate",
    "transport_power_imponibile_eur",
    "uc3_qty",
    "uc3_unit_rate",
    "uc3_imponibile_eur",
    "uc6_fixed_qty",
    "uc6_fixed_unit_rate",
    "uc6_fixed_imponibile_eur",
    "uc6_variable_qty",
    "uc6_variable_unit_rate",
    "uc6_variable_imponibile_eur",
    "arim_qty",
    "arim_unit_rate",
    "arim_imponibile_eur",
    "asos_qty",
    "asos_unit_rate",
    "asos_imponibile_eur",
    "excise_qty",
    "excise_unit_rate",
    "excise_imponibile_eur",
    "notes",
]

NUMERIC_COLUMNS = [
    "consumption_kwh",
    "consumption_f1_kwh",
    "consumption_f2_kwh",
    "consumption_f3_kwh",
    "committed_power_kw",
    "total_amount_eur",
    "invoice_total_eur",
    "bonus_eur",
    "energy_cost_eur",
    "transport_eur",
    "system_charges_eur",
    "taxes_eur",
    "vat_eur",
    "tv_license_eur",
    "energy_qty",
    "energy_imponibile_eur",
    "losses_qty",
    "losses_imponibile_eur",
    "dispbt_qty",
    "dispbt_imponibile_eur",
    "commercialization_qty",
    "commercialization_imponibile_eur",
    "capacity_market_qty",
    "capacity_market_imponibile_eur",
    "dispatching_qty",
    "dispatching_imponibile_eur",
    "transport_energy_qty",
    "transport_energy_imponibile_eur",
    "transport_fixed_qty",
    "transport_fixed_imponibile_eur",
    "transport_power_qty",
    "transport_power_imponibile_eur",
    "uc3_qty",
    "uc3_imponibile_eur",
    "uc6_fixed_qty",
    "uc6_fixed_imponibile_eur",
    "uc6_variable_qty",
    "uc6_variable_imponibile_eur",
    "arim_qty",
    "arim_imponibile_eur",
    "asos_qty",
    "asos_imponibile_eur",
    "excise_qty",
    "excise_imponibile_eur",
]


@dataclass
class BillRecord:
    source_file: str
    supplier_template: str = ""
    supplier_name: str = ""
    invoice_number: str = ""
    invoice_date: str = ""
    due_date: str = ""
    customer_name: str = ""
    supply_address: str = ""
    pod_code: str = ""
    tariff_code: str = ""
    billing_period_start: str = ""
    billing_period_end: str = ""
    consumption_kwh: str = ""
    consumption_f1_kwh: str = ""
    consumption_f2_kwh: str = ""
    consumption_f3_kwh: str = ""
    committed_power_kw: str = ""
    total_amount_eur: str = ""
    invoice_total_eur: str = ""
    bonus_eur: str = ""
    energy_cost_eur: str = ""
    transport_eur: str = ""
    system_charges_eur: str = ""
    taxes_eur: str = ""
    vat_eur: str = ""
    tv_license_eur: str = ""
    energy_qty: str = ""
    energy_unit_rate: str = ""
    energy_imponibile_eur: str = ""
    losses_qty: str = ""
    losses_unit_rate: str = ""
    losses_imponibile_eur: str = ""
    dispbt_qty: str = ""
    dispbt_unit_rate: str = ""
    dispbt_imponibile_eur: str = ""
    commercialization_qty: str = ""
    commercialization_unit_rate: str = ""
    commercialization_imponibile_eur: str = ""
    capacity_market_qty: str = ""
    capacity_market_unit_rate: str = ""
    capacity_market_imponibile_eur: str = ""
    dispatching_qty: str = ""
    dispatching_unit_rate: str = ""
    dispatching_imponibile_eur: str = ""
    transport_energy_qty: str = ""
    transport_energy_unit_rate: str = ""
    transport_energy_imponibile_eur: str = ""
    transport_fixed_qty: str = ""
    transport_fixed_unit_rate: str = ""
    transport_fixed_imponibile_eur: str = ""
    transport_power_qty: str = ""
    transport_power_unit_rate: str = ""
    transport_power_imponibile_eur: str = ""
    uc3_qty: str = ""
    uc3_unit_rate: str = ""
    uc3_imponibile_eur: str = ""
    uc6_fixed_qty: str = ""
    uc6_fixed_unit_rate: str = ""
    uc6_fixed_imponibile_eur: str = ""
    uc6_variable_qty: str = ""
    uc6_variable_unit_rate: str = ""
    uc6_variable_imponibile_eur: str = ""
    arim_qty: str = ""
    arim_unit_rate: str = ""
    arim_imponibile_eur: str = ""
    asos_qty: str = ""
    asos_unit_rate: str = ""
    asos_imponibile_eur: str = ""
    excise_qty: str = ""
    excise_unit_rate: str = ""
    excise_imponibile_eur: str = ""
    notes: str = ""


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.replace("\x00", " ")).strip()


def slug_text(value: str) -> str:
    lowered = value.lower()
    replacements = str.maketrans("àèéìíîòóùú", "aeeiiioouu")
    return normalize_text(lowered.translate(replacements))


def extract_pdf_text(pdf_path: Path) -> str:
    reader = PdfReader(str(pdf_path))
    pages: list[str] = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\n".join(pages)


def extract_lines(text: str) -> list[str]:
    return [normalize_text(line) for line in text.splitlines() if normalize_text(line)]


def parse_date(value: str) -> str:
    cleaned = value.strip(" .,:;")
    cleaned = re.sub(r"(\d)([A-Za-z])", r"\1 \2", cleaned)
    cleaned = re.sub(r"([A-Za-z])(\d)", r"\1 \2", cleaned)
    for italian, english in ITALIAN_MONTHS.items():
        cleaned = re.sub(italian, english, cleaned, flags=re.IGNORECASE)
    try:
        parsed = date_parser.parse(cleaned, dayfirst=True, fuzzy=True)
    except (ValueError, OverflowError):
        return ""
    return parsed.date().isoformat()


def parse_decimal(value: str) -> str:
    cleaned = value.strip().replace("−", "-")
    cleaned = re.sub(r"[^\d,.\-]", "", cleaned)
    if not cleaned:
        return ""
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "." in cleaned and cleaned.count(".") >= 1:
        parts = cleaned.split(".")
        if all(part.isdigit() for part in parts) and all(len(part) == 3 for part in parts[1:]):
            cleaned = "".join(parts)
    try:
        return f"{float(cleaned):.2f}"
    except ValueError:
        return ""


def parse_float_num(value: str) -> float | None:
    cleaned = value.strip().replace("−", "-")
    cleaned = re.sub(r"[^\d,.\-]", "", cleaned)
    if not cleaned:
        return None
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif "," in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "." in cleaned and cleaned.count(".") >= 1:
        parts = cleaned.split(".")
        if all(part.isdigit() for part in parts) and all(len(part) == 3 for part in parts[1:]):
            cleaned = "".join(parts)
    try:
        return float(cleaned)
    except ValueError:
        return None


def parse_number(value: str) -> str:
    parsed = parse_decimal(value)
    if not parsed:
        return ""
    numeric = float(parsed)
    return str(int(numeric)) if numeric.is_integer() else parsed


def extract_with_patterns(text: str, patterns: list[tuple[str, str]]) -> str:
    for pattern, kind in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
        if not match:
            continue
        value = match.group(1)
        return extract_by_kind(value, kind) if kind != "text" else normalize_text(value)
    return ""


def find_label_value(lines: list[str], labels: Iterable[str], kind: str) -> str:
    for idx, line in enumerate(lines):
        slug = slug_text(line)
        for label in labels:
            s_label = slug_text(label)
            if s_label not in slug:
                continue
            direct = try_extract_after_label(line, label, kind)
            if direct:
                return direct
            if idx + 1 < len(lines):
                next_line = lines[idx + 1]
                if label_is_too_generic(next_line):
                    continue
                candidate = extract_by_kind(next_line, kind)
                if candidate:
                    return candidate
                return next_line
    return ""


def try_extract_after_label(line: str, label: str, kind: str) -> str:
    pattern = re.compile(re.escape(label), re.IGNORECASE)
    match = pattern.search(line)
    if not match:
        return ""
    tail = line[match.end() :]
    tail = re.sub(r"^[\s:.-]+", "", tail)
    return extract_by_kind(tail, kind) or normalize_text(tail)


def label_is_too_generic(value: str) -> bool:
    slug = slug_text(value)
    return any(hint in slug for hint in DATE_HINTS)


def extract_by_kind(value: str, kind: str) -> str:
    if kind == "date":
        match = re.search(
            r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{1,2}\s+[A-Za-z]+\s+\d{2,4})",
            value,
            re.IGNORECASE,
        )
        return parse_date(match.group(1)) if match else ""
    if kind == "money":
        match = re.search(r"(-?\d[\d.\s]*,\d{2}|-?\d[\d,.\s]*\.\d{2})", value)
        return parse_decimal(match.group(1)) if match else ""
    if kind == "code":
        match = re.search(r"\b([A-Z]{2}\d[A-Z0-9]{11,14}|IT\d{14}[A-Z0-9]{0,4}|[A-Z0-9][A-Z0-9/\-]{5,})\b", value)
        return match.group(1) if match else normalize_text(value)
    if kind == "number":
        match = re.search(r"(-?\d[\d.\s]*,\d+|-?\d[\d,.\s]*\.\d+|-?\d+)", value)
        return parse_number(match.group(1)) if match else ""
    return normalize_text(value)


def find_period(lines: list[str], raw_text: str) -> tuple[str, str]:
    combined = "\n".join(lines)
    patterns = [
        r"dal\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s+al\s+(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
        r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*[-–]\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})",
    ]
    for label in PERIOD_LABELS:
        for line in lines:
            if slug_text(label) in slug_text(line):
                for pattern in patterns:
                    match = re.search(pattern, line, re.IGNORECASE)
                    if match:
                        return parse_date(match.group(1)), parse_date(match.group(2))
    for pattern in patterns:
        match = re.search(pattern, combined, re.IGNORECASE)
        if match:
            return parse_date(match.group(1)), parse_date(match.group(2))
    # Fallback to month span in document.
    month_matches = re.findall(r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", raw_text)
    if len(month_matches) >= 2:
        return parse_date(month_matches[0]), parse_date(month_matches[1])
    return "", ""


def find_consumption(lines: list[str], raw_text: str) -> str:
    candidates: list[float] = []
    patterns = [
        r"consum[oi]\w*\D{0,12}(\d[\d.\s]*,\d+|\d[\d.\s]*)\s*kwh",
        r"energia attiva\D{0,12}(\d[\d.\s]*,\d+|\d[\d.\s]*)\s*kwh",
        r"totale kwh\D{0,12}(\d[\d.\s]*,\d+|\d[\d.\s]*)",
    ]
    for line in lines:
        lowered = slug_text(line)
        if "kwh" not in lowered:
            continue
        for pattern in patterns:
            match = re.search(pattern, lowered, re.IGNORECASE)
            if match:
                parsed = parse_decimal(match.group(1))
                if parsed:
                    candidates.append(float(parsed))
    if not candidates:
        for match in re.finditer(r"(\d[\d.\s]*,\d+|\d[\d.\s]*)\s*kwh", raw_text, re.IGNORECASE):
            parsed = parse_decimal(match.group(1))
            if parsed:
                candidates.append(float(parsed))
    if not candidates:
        return ""
    return str(int(max(candidates))) if max(candidates).is_integer() else f"{max(candidates):.2f}"


def find_committed_power(lines: list[str], raw_text: str) -> str:
    patterns = [
        r"potenza impegnata\D{0,10}(\d[\d.,]*)\s*kw",
        r"potenza disponibile\D{0,10}(\d[\d.,]*)\s*kw",
    ]
    haystack = "\n".join(lines) + "\n" + raw_text
    for pattern in patterns:
        match = re.search(pattern, slug_text(haystack), re.IGNORECASE)
        if match:
            return parse_number(match.group(1))
    return ""


def infer_supplier(lines: list[str], pdf_path: Path) -> str:
    for line in lines[:6]:
        if "octopus energy" in slug_text(line):
            return normalize_text(line[:120])
    joined = " ".join(lines[:5]).strip()
    if joined:
        pieces = re.split(
            r"\b(fattura|numero fattura|data emissione|cliente|pod|scadenza)\b",
            joined,
            flags=re.IGNORECASE,
        )
        candidate = normalize_text(pieces[0])
        if candidate:
            return candidate[:120]
    if lines:
        return lines[0][:120]
    return pdf_path.stem


def infer_supplier_template(raw_text: str, lines: list[str]) -> str:
    haystack = slug_text("\n".join(lines[:40]) + "\n" + raw_text[:5000])
    if "octopus energy" in haystack:
        return "octopus"
    if "acea energia" in haystack or "acea" in haystack:
        if "periodo di conguaglio" in haystack or "periodica + conguaglio" in haystack:
            return "acea_conguaglio"
        return "acea_standard"
    return "generic"


def build_regex_overrides(raw_text: str, lines: list[str]) -> dict[str, str]:
    overrides: dict[str, str] = {}
    patterns = {
        "supplier_name": [
            (r"^(Octopus Energy Italia Srl)", "text"),
            (r"^(Octopus Energy[^\n]+)", "text"),
        ],
        "invoice_number": [
            (r"NUMERO FATTURA[^\n]*\n\s*([A-Z0-9\-]+)", "text"),
            (r"Fattura\s*n[.\s:]*([A-Z0-9\-\/]+)", "text"),
        ],
        "invoice_date": [
            (r"DATA FATTURA:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})", "date"),
            (r"DATA EMISSIONE:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})", "date"),
        ],
        "due_date": [
            (r"Entro il\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})", "date"),
            (r"SCADENZA:\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})", "date"),
        ],
        "customer_name": [
            (r"I tuoi dati\s*\n\s*([^\n]+)", "text"),
            (r"Cliente:\s*([^\n]+)", "text"),
        ],
        "supply_address": [
            (r"INDIRIZZO DI FORNITURA:\s*\n\s*([^\n]+)", "text"),
            (r"INDIRIZZO DI FATTURAZIONE:\s*\n\s*([^\n]+)", "text"),
        ],
        "pod_code": [
            (r"CODICE POD:\s*\n\s*([A-Z0-9]+)", "code"),
            (r"\b(IT\d{3}E\d{8,})\b", "code"),
        ],
        "tariff_code": [
            (r"NOME OFFERTA:\s*([^\n]+)", "text"),
            (r"OFFERTA:\s*([^\n]+)", "text"),
        ],
        "consumption_kwh": [
            (r"CONSUMO FATTURATO:\s*([0-9.,]+)\s*kWh", "number"),
        ],
        "committed_power_kw": [
            (r"POTENZA IMPEGNATA:\s*\n\s*([0-9.,]+)\s*kW", "number"),
            (r"Potenza impegnata\s*([0-9.,]+)\s*kW", "number"),
        ],
        "total_amount_eur": [
            (r"TOTALE DA PAGARE\s*\n\s*([0-9.,\-]+)\s*€", "money"),
            (r"Questo mese dovrai pagare\s*\n\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "invoice_total_eur": [
            (r"TOTALE BOLLETTA\s*\n\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "bonus_eur": [
            (r"Bonus applicati\s*\n\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "taxes_eur": [
            (r"Accise e IVA\s*\n\s*([0-9.,\-]+)\s*€", "money"),
            (r"Totale Imposte\s*\n\s*([0-9.,\-]+)\s*€", "money"),
            (r"IMPOSTE\s*\n?\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "vat_eur": [
            (r"DI CUI IVA\s*\n\s*([0-9.,\-]+)\s*€", "money"),
            (r"IVA vendite 10%:\s*[0-9.,\-]+\s*€\s*x\s*10%\s*=\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "transport_eur": [
            (r"Spesa per il trasporto e la gestione del contatore[\s\S]*?TOTALE\s*\n\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "system_charges_eur": [
            (r"Spesa per oneri di sistema[\s\S]*?TOTALE\s*\n\s*([0-9.,\-]+)\s*€", "money"),
            (r"Totale oneri di sistema\s*\n\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "tv_license_eur": [
            (r"Canone di abbonamento alla televisione[^\n]*\n\s*([0-9.,\-]+)\s*€", "money"),
        ],
    }

    for field, field_patterns in patterns.items():
        extracted = extract_with_patterns(raw_text, field_patterns)
        if extracted:
            overrides[field] = extracted

    network_system_amounts = [
        parse_decimal(match.group(1))
        for match in re.finditer(
            r"di cui spesa per la rete e gli oneri generali di sistema\s*\n\s*[0-9.,]+\s*[^\n]*\n\s*([0-9.,\-]+)\s*€",
            raw_text,
            re.IGNORECASE,
        )
    ]
    sale_energy_amounts = [
        parse_decimal(match.group(1))
        for match in re.finditer(
            r"di cui spesa per vendita energia elettrica\s*\n\s*[0-9.,]+\s*[^\n]*\n\s*([0-9.,\-]+)\s*€",
            raw_text,
            re.IGNORECASE,
        )
    ]
    if network_system_amounts:
        total = sum(float(value) for value in network_system_amounts if value)
        overrides["system_charges_eur"] = f"{total:.2f}"
    if sale_energy_amounts:
        total = sum(float(value) for value in sale_energy_amounts if value)
        overrides["energy_cost_eur"] = f"{total:.2f}"

    if not overrides.get("customer_name"):
        for idx, line in enumerate(lines):
            if slug_text(line) == "i tuoi dati" and idx + 1 < len(lines):
                overrides["customer_name"] = lines[idx + 1]
                break

    return overrides


def build_octopus_regex_overrides(raw_text: str, lines: list[str]) -> dict[str, str]:
    return build_regex_overrides(raw_text, lines)


def build_generic_regex_overrides(raw_text: str, lines: list[str]) -> dict[str, str]:
    overrides: dict[str, str] = {}
    patterns = {
        "invoice_number": [
            (r"(?:NUMERO FATTURA|FATTURA N\.?|NUMERO DOCUMENTO)[^\n:]*[: ]\s*([A-Z0-9\-\/]+)", "text"),
        ],
        "invoice_date": [
            (r"(?:DATA FATTURA|DATA EMISSIONE|EMESSA IL)[: ]\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})", "date"),
        ],
        "due_date": [
            (r"(?:SCADENZA|DATA SCADENZA|ENTRO IL)[: ]\s*([0-9]{1,2}/[0-9]{1,2}/[0-9]{2,4})", "date"),
        ],
        "pod_code": [
            (r"\b(IT\d{3}E\d{8,})\b", "code"),
        ],
        "committed_power_kw": [
            (r"(?:POTENZA IMPEGNATA|POTENZA DISPONIBILE)[: ]\s*([0-9.,]+)\s*kW", "number"),
        ],
        "consumption_kwh": [
            (r"(?:CONSUMO FATTURATO|CONSUMO TOTALE|ENERGIA ATTIVA)[: ]\s*([0-9.,]+)\s*kWh", "number"),
        ],
        "total_amount_eur": [
            (r"(?:TOTALE DA PAGARE|IMPORTO DA PAGARE|TOTALE FATTURA)[: ]?\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "invoice_total_eur": [
            (r"TOTALE BOLLETTA[: ]?\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "taxes_eur": [
            (r"(?:TOTALE IMPOSTE|ACCISE E IVA|IMPOSTE)[: ]?\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "vat_eur": [
            (r"(?:DI CUI IVA|IVA)[: ]?\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "transport_eur": [
            (r"(?:SPESA PER IL TRASPORTO E LA GESTIONE DEL CONTATORE|TRASPORTO E GESTIONE DEL CONTATORE)[: ]?\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "system_charges_eur": [
            (r"(?:TOTALE ONERI DI SISTEMA|SPESA PER ONERI DI SISTEMA|ONERI DI SISTEMA)[: ]?\s*([0-9.,\-]+)\s*€", "money"),
        ],
        "tv_license_eur": [
            (r"(?:CANONE RAI|CANONE TV|CANONE DI ABBONAMENTO ALLA TELEVISIONE)[^\n]*[: ]?\s*([0-9.,\-]+)\s*€", "money"),
        ],
    }
    for field, field_patterns in patterns.items():
        extracted = extract_with_patterns(raw_text, field_patterns)
        if extracted:
            overrides[field] = extracted
    return overrides


def build_acea_regex_overrides(raw_text: str, lines: list[str]) -> dict[str, str]:
    overrides = build_generic_regex_overrides(raw_text, lines)
    patterns = {
        "supplier_name": [
            (r"(Acea Energia S\.p\.A\.)", "text"),
            (r"^(ACEA[^\n]+)", "text"),
            (r"^(Acea[^\n]+)", "text"),
        ],
        "invoice_number": [
            (r"BOLLETTA PER LA FORNITURA\s+DI ENERGIA ELETTRICAn\.\s*([A-Z0-9\-]+)\s+del", "text"),
        ],
        "invoice_date": [
            (r"BOLLETTA PER LA FORNITURA\s+DI ENERGIA ELETTRICAn\.\s*[A-Z0-9\-]+\s+del\s+([0-9]{2}/[0-9]{2}/[0-9]{4})", "date"),
        ],
        "tariff_code": [
            (r"OFFERTA\s+([^\n]+)", "text"),
            (r"(?:NOME OFFERTA|TIPO OFFERTA|DENOMINAZIONE OFFERTA)[: ]\s*([^\n]+)", "text"),
        ],
        "supply_address": [
            (r"INDIRIZZO FORNITURA\s+([^\n]+(?:\n[^\n]+)?)", "text"),
            (r"(?:INDIRIZZO DI FORNITURA|UBICAZIONE FORNITURA)[: ]\s*([^\n]+)", "text"),
        ],
        "customer_name": [
            (r"INTESTATARIO CONTRATTO\s+([^\n]+)", "text"),
            (r"(?:INTESTATARIO|CLIENTE|NOMINATIVO)[: ]\s*([^\n]+)", "text"),
        ],
        "due_date": [
            (r"ENTRO IL\s+([0-9]{2}/[0-9]{2}/[0-9]{4})", "date"),
        ],
        "billing_period_start": [
            (r"PERIODO DI FATTURAZIONE:\s*([0-9]{2}\s+[A-Z]+?\s+[0-9]{4})\s*-\s*([0-9]{2}\s+[A-Z]+?\s+[0-9]{4})", "text"),
        ],
        "total_amount_eur": [
            (r"TOTALE DA PAGARE\s+([0-9.]+,\d{2})\s+euro", "money"),
            (r"TOTALE DA PAGARE\s+([0-9.]+,\d{2})\s+EURO", "money"),
        ],
        "invoice_total_eur": [
            (r"TOTALE BOLLETTA\s+([0-9.]+,\d{2})\s+EURO", "money"),
        ],
        "energy_cost_eur": [
            (r"SPESA PER LA MATERIA ENERGIA\s+([0-9.]+,\d{2})\s+EURO", "money"),
            (r"SPESA PER LA MATERIA ENERGIA TOTALE\s+([0-9.]+,\d{2})\s+€", "money"),
        ],
        "transport_eur": [
            (r"SPESA PER IL TRASPORTO E LA GESTIONE CONTATORE TOTALE\s+([0-9.]+,\d{2})\s+€", "money"),
            (r"SPESA PER IL TRASPORTO E LA GESTIONE DEL CONTATORE\s+([0-9.]+,\d{2})\s+EURO", "money"),
        ],
        "system_charges_eur": [
            (r"SPESA PER ONERI DI SISTEMA TOTALE\s+(?:\d+\s+)?([0-9.]+,\d{2})\s+€", "money"),
            (r"SPESA PER ONERI DI SISTEMA\*+\s+([0-9.]+,\d{2})\s+EURO", "money"),
        ],
        "taxes_eur": [
            (r"TOTALE IMPOSTE E IVA(?:\s+\d+)?\s+([0-9.]+,\d{2})\s+EURO", "money"),
            (r"TOTALE IMPOSTE E IVA\*+\s+([0-9.]+,\d{2})\s+EURO", "money"),
            (r"TOTALE\s+([0-9.]+,\d{2})\s+euro", "money"),
        ],
        "vat_eur": [
            (r"IVA 10% immediata su imponibile di [0-9.]+,\d{2}\s+EURO\s+([0-9.]+,\d{2})", "money"),
            (r"IVA 10% immediata su imponibile di [0-9.]+,\d{2}\s+euro\s+([0-9.]+,\d{2})", "money"),
        ],
        "tv_license_eur": [
            (r"CANONE DI ABBONAMENTO TV\s+([0-9.]+,\d{2})\s+EURO", "money"),
            (r"CANONE DI ABBONAMENTO ALLA TELEVISIONE PER USO PRIVATO TOTALE\s+([0-9.]+,\d{2})\s+€", "money"),
            (r"Canone TV[: ]\s*([0-9.]+,\d{2})\s+EURO", "money"),
        ],
    }
    for field, field_patterns in patterns.items():
        extracted = extract_with_patterns(raw_text, field_patterns)
        if extracted:
            overrides[field] = extracted
    period_match = re.search(
        r"PERIODO DI FATTURAZIONE:\s*([0-9]{2}\s+[A-Z]+?\s+[0-9]{4})\s*-\s*([0-9]{2}\s+[A-Z]+?\s+[0-9]{4})",
        raw_text,
        re.IGNORECASE,
    )
    if period_match:
        overrides["billing_period_start"] = parse_date(period_match.group(1))
        overrides["billing_period_end"] = parse_date(period_match.group(2))
    return overrides


def extract_detail_values(raw_text: str, label_patterns: list[str]) -> tuple[str, str, str]:
    label_block = "(?:" + "|".join(label_patterns) + ")"
    pattern = re.compile(
        label_block
        + r"\s*\n\s*([0-9.,]+)\s*\n\s*([0-9.,]+(?:\s*€/[\w]+)?)\s*\n\s*([0-9.,\-]+)\s*€",
        re.IGNORECASE | re.MULTILINE,
    )
    match = pattern.search(raw_text)
    if not match:
        return "", "", ""
    qty = parse_number(match.group(1))
    unit_rate = normalize_text(match.group(2))
    imponibile = parse_decimal(match.group(3))
    return qty, unit_rate, imponibile


def decimal_to_str(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else f"{value:.2f}"


def normalize_unit_rate(value: float, unit: str) -> str:
    unit = normalize_text(unit)
    if not unit:
        return f"{value:.6f}"
    return f"{value:.6f} {unit}"


def sum_amounts(values: list[str]) -> str:
    total = sum(float(v) for v in values if v)
    return f"{total:.2f}" if values else ""


def extract_section(raw_text: str, start_label: str, end_label: str | None = None) -> str:
    start = re.search(start_label, raw_text, re.IGNORECASE)
    if not start:
        return ""
    tail = raw_text[start.end():]
    if end_label:
        end = re.search(end_label, tail, re.IGNORECASE)
        if end:
            tail = tail[: end.start()]
    return tail


def aggregate_acea_rows(block: str) -> tuple[str, str, str]:
    row_patterns = [
        re.compile(
            r"^(?:\d{2}/\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}\s+)?(?:F\d|SCAGLIONE\s+\d+|\d+)?\s*"
            r"(€/[A-Za-z/]+|€/[\w]+|euro/[A-Za-z/]+|Euro/[A-Za-z/]+)\s+([\-−]?\d+[.,]\d+)\s+"
            r"(?:kWh|Kwh|kW|gg)?\s*([\d.]+,\d+|[\d.]+)\s+([\-−]?\d+,\d+|\d+\.\d+)$",
            re.IGNORECASE,
        ),
        re.compile(
            r"^(?:\d{2}/\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}\s+)?(?:F\d|SCAGLIONE\s+\d+|\d+)?\s*"
            r"(€/[A-Za-z/]+|€/[\w]+|euro/[A-Za-z/]+|Euro/[A-Za-z/]+)\s+([\-−]?\d+[.,]\d+)\s+"
            r"([\d.]+,\d+|[\d.]+)\s*(?:kWh|Kwh|kW|gg)?\s+([\-−]?\d+,\d+|\d+\.\d+)$",
            re.IGNORECASE,
        ),
    ]
    quantities: list[float] = []
    imponibili: list[float] = []
    unit_rates: list[tuple[float, float]] = []
    preferred_unit = ""
    for line in [normalize_text(x) for x in block.splitlines() if normalize_text(x)]:
        match = None
        for candidate in row_patterns:
            match = candidate.match(line)
            if match:
                break
        if not match:
            continue
        preferred_unit = preferred_unit or match.group(1).replace("Euro", "euro")
        unit_rate = parse_float_num(match.group(2))
        qty = parse_decimal(match.group(3))
        imponibile = parse_decimal(match.group(4))
        if not qty or not imponibile:
            continue
        qty_f = float(qty)
        imp_f = float(imponibile)
        quantities.append(qty_f)
        imponibili.append(imp_f)
        if unit_rate is not None:
            unit_rates.append((unit_rate, qty_f))
    if not quantities or not imponibili:
        return "", "", ""
    qty_total = sum(quantities)
    imponibile_total = sum(imponibili)
    avg_rate = ""
    if unit_rates and qty_total:
        weighted = sum(rate * qty for rate, qty in unit_rates) / qty_total
        avg_rate = normalize_unit_rate(weighted, preferred_unit)
    return decimal_to_str(qty_total), avg_rate, f"{imponibile_total:.2f}"


def aggregate_acea_fixed_rows(block: str) -> tuple[str, str, str]:
    row_patterns = [
        re.compile(
            r"^(?:\d{2}/\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}\s+)?(€/[A-Za-z/]+|€/[\w]+|Euro/[A-Za-z/]+|euro/[A-Za-z/]+)\s+([\-−]?\d+[.,]\d+)\s+"
            r"(?:gg|kW)?\s*([\d.]+,\d+|[\d.]+)\s+([\-−]?\d+,\d+|\d+\.\d+)$",
            re.IGNORECASE,
        ),
        re.compile(
            r"^(?:\d{2}/\d{2}/\d{4}\s+\d{2}/\d{2}/\d{4}\s+)?(€/[A-Za-z/]+|€/[\w]+|Euro/[A-Za-z/]+|euro/[A-Za-z/]+)\s+([\-−]?\d+[.,]\d+)\s+"
            r"([\d.]+,\d+|[\d.]+)\s*(?:gg|kW)?\s+([\-−]?\d+,\d+|\d+\.\d+)$",
            re.IGNORECASE,
        ),
    ]
    quantities: list[float] = []
    imponibili: list[float] = []
    unit_rates: list[tuple[float, float]] = []
    preferred_unit = ""
    for line in [normalize_text(x) for x in block.splitlines() if normalize_text(x)]:
        match = None
        for candidate in row_patterns:
            match = candidate.match(line)
            if match:
                break
        if not match:
            continue
        preferred_unit = match.group(1).replace("Euro", "euro")
        unit_rate = parse_float_num(match.group(2))
        qty = parse_decimal(match.group(3))
        imponibile = parse_decimal(match.group(4))
        if not qty or not imponibile:
            continue
        qty_f = float(qty)
        imp_f = float(imponibile)
        quantities.append(qty_f)
        imponibili.append(imp_f)
        if unit_rate is not None:
            unit_rates.append((unit_rate, qty_f))
    if not quantities or not imponibili:
        return "", "", ""
    qty_total = sum(quantities)
    imponibile_total = sum(imponibili)
    avg_rate = ""
    if unit_rates and qty_total:
        weighted = sum(rate * qty for rate, qty in unit_rates) / qty_total
        avg_rate = normalize_unit_rate(weighted, preferred_unit)
    return decimal_to_str(qty_total), avg_rate, f"{imponibile_total:.2f}"


def extract_fascia_consumptions(raw_text: str) -> tuple[str, str, str]:
    match = re.search(
        r"Letture.*?(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(\d+)\s+(\d+)\s+Rilevata\s+(\d{2}/\d{2}/\d{4})\s+(\d+)\s+(\d+)\s+(\d+)\s+Rilevata",
        raw_text,
        re.IGNORECASE | re.DOTALL,
    )
    if not match:
        return "", "", ""
    f1 = str(int(match.group(2)) - int(match.group(6)))
    f2 = str(int(match.group(3)) - int(match.group(7)))
    f3 = str(int(match.group(4)) - int(match.group(8)))
    return f1, f2, f3


def extract_acea_fascia_consumptions(raw_text: str) -> tuple[str, str, str]:
    totals = {"F1": 0.0, "F2": 0.0, "F3": 0.0}
    letture_block = extract_section(raw_text, r"PROSPETTO LETTURE E CONSUMI", r"DETTAGLIO CONSUMO FATTURATO")
    for line in [normalize_text(x) for x in letture_block.splitlines() if normalize_text(x)]:
        simple = re.search(r"\b(F[123])\b.*?\b(\d+)\s*kWh\b", line, re.IGNORECASE)
        if simple:
            qty = parse_float_num(simple.group(2))
            if qty is not None:
                totals[simple.group(1).upper()] += qty
    if any(totals.values()):
        return (
            decimal_to_str(totals["F1"]),
            decimal_to_str(totals["F2"]),
            decimal_to_str(totals["F3"]),
        )
    return "", "", ""


def build_detail_component_overrides(raw_text: str) -> dict[str, str]:
    components = {
        "energy": [r"Energia"],
        "losses": [r"Perdite"],
        "dispbt": [r"Componente DISPbt"],
        "commercialization": [r"Corrispettivo di commercializzazione", r"Corrispettivo commercializzazione"],
        "capacity_market": [r"Corrispettivo mercato capacit(?:à|a)"],
        "dispatching": [r"Dispacciamento"],
        "transport_energy": [r"Trasporto quota energia"],
        "transport_fixed": [r"Trasporto quota fissa"],
        "transport_power": [r"Trasporto quota potenza"],
        "uc3": [r"UC3[^\n]*(?:\n\s*[^\n]*)?"],
        "uc6_fixed": [r"UC6[^\n]*\n\s*\(fisso\)"],
        "uc6_variable": [r"UC6[^\n]*\n\s*\(variabile\)"],
        "arim": [r"Componente ARIM[^\n]*(?:\n[^\n]*)?"],
        "asos": [r"Componente ASOS[^\n]*(?:\n[^\n]*)?"],
        "excise": [r"Imposta erariale[^\n]*"],
    }

    overrides: dict[str, str] = {}
    for prefix, patterns in components.items():
        qty, unit_rate, imponibile = extract_detail_values(raw_text, patterns)
        overrides[f"{prefix}_qty"] = qty
        overrides[f"{prefix}_unit_rate"] = unit_rate
        overrides[f"{prefix}_imponibile_eur"] = imponibile

    return overrides


def build_generic_detail_component_overrides(raw_text: str) -> dict[str, str]:
    return {}


def extract_generic_fascia_consumptions(raw_text: str) -> tuple[str, str, str]:
    return "", "", ""


def build_acea_detail_component_overrides(raw_text: str) -> dict[str, str]:
    overrides: dict[str, str] = {}

    component_blocks = {
        "energy": extract_section(raw_text, r"COMPONENTE ENERGIA", r"DISPACCIAMENTO"),
        "dispatching": extract_section(
            raw_text,
            r"DISPACCIAMENTO\s+(?:DA A|DAL AL)\s+FASCIA(?:\s+UNITA'\s+DI\s+MISURA)?\s+PREZZO UNITARIO\s+QUANTITÀ EURO",
            r"PERDITE DI RETE - DISPACCIAMENTO",
        ),
        "losses_disp": extract_section(raw_text, r"PERDITE DI RETE - DISPACCIAMENTO", r"PERDITE DI RETE - ENERGIA"),
        "losses_energy": extract_section(raw_text, r"PERDITE DI RETE - ENERGIA", r"QUOTA FISSA"),
        "dispbt": extract_section(raw_text, r"COMP\.DI DISPACCIAMENTO \(PARTE FISSA\)", r"COMPONENTE FISSA QV1 FLEX"),
        "commercialization": extract_section(raw_text, r"COMPONENTE FISSA QV1 FLEX", r"SPESA PER IL TRASPORTO E LA GESTIONE CONTATORE"),
        "transport_energy": extract_section(raw_text, r"QUOTA VARIABILE", r"QUOTA FISSA"),
        "transport_fixed": extract_section(raw_text, r"QUOTA FISSA\s+QUOTA FISSA", r"QUOTA POTENZA"),
        "transport_power": extract_section(raw_text, r"QUOTA POTENZA\s+QUOTA POTENZA", r"UC6 QUOTA POTENZA"),
        "uc6_fixed": extract_section(raw_text, r"UC6 QUOTA POTENZA", r"SPESA PER ONERI DI SISTEMA TOTALE"),
        "arim": extract_section(raw_text, r"COMPONENTE ARIM", r"COMPONENTE ASOS"),
        "asos": extract_section(raw_text, r"COMPONENTE ASOS", r"IMPOSTE E IVA TOTALE"),
        "excise": extract_section(
            raw_text,
            r"IMPOSTA ERARIALE\s+(?:DA A|DAL AL)(?:\s+UNITA'\s+DI\s+MISURA)?\s+PREZZO UNITARIO\s+QUANTITÀ EURO",
            r"IVA 10%",
        ),
    }

    qty, unit_rate, imponibile = aggregate_acea_rows(component_blocks["energy"])
    overrides["energy_qty"] = qty
    overrides["energy_unit_rate"] = unit_rate
    overrides["energy_imponibile_eur"] = imponibile

    qty_loss_e, rate_loss_e, imp_loss_e = aggregate_acea_rows(component_blocks["losses_energy"])
    qty_loss_d, rate_loss_d, imp_loss_d = aggregate_acea_rows(component_blocks["losses_disp"])
    if qty_loss_e or qty_loss_d:
        q_total = sum(float(x) for x in [qty_loss_e, qty_loss_d] if x)
        i_total = sum(float(x) for x in [imp_loss_e, imp_loss_d] if x)
        weighted_parts: list[tuple[float, float]] = []
        if qty_loss_e and rate_loss_e:
            parsed = parse_float_num(rate_loss_e)
            if parsed is not None:
                weighted_parts.append((parsed, float(qty_loss_e)))
        if qty_loss_d and rate_loss_d:
            parsed = parse_float_num(rate_loss_d)
            if parsed is not None:
                weighted_parts.append((parsed, float(qty_loss_d)))
        avg = ""
        if weighted_parts and q_total:
            avg = normalize_unit_rate(sum(r * q for r, q in weighted_parts) / q_total, "euro/kWh")
        overrides["losses_qty"] = decimal_to_str(q_total)
        overrides["losses_unit_rate"] = avg
        overrides["losses_imponibile_eur"] = f"{i_total:.2f}"

    for prefix, block, fixed in [
        ("dispatching", component_blocks["dispatching"], False),
        ("dispbt", component_blocks["dispbt"], True),
        ("commercialization", component_blocks["commercialization"], True),
        ("transport_energy", component_blocks["transport_energy"], False),
        ("transport_fixed", component_blocks["transport_fixed"], True),
        ("transport_power", component_blocks["transport_power"], True),
        ("uc6_fixed", component_blocks["uc6_fixed"], True),
        ("arim", component_blocks["arim"], False),
        ("asos", component_blocks["asos"], False),
        ("excise", component_blocks["excise"], False),
    ]:
        qty, unit_rate, imponibile = (
            aggregate_acea_fixed_rows(block) if fixed else aggregate_acea_rows(block)
        )
        overrides[f"{prefix}_qty"] = qty
        overrides[f"{prefix}_unit_rate"] = unit_rate
        overrides[f"{prefix}_imponibile_eur"] = imponibile

    return overrides


def clean_acea_record(record: BillRecord) -> None:
    if record.supplier_name and "acea.it" in record.supplier_name.lower():
        record.supplier_name = "Acea Energia S.p.A."
    if record.tv_license_eur and not re.fullmatch(r"\d+\.\d{2}", record.tv_license_eur):
        record.tv_license_eur = ""


def clean_octopus_record(record: BillRecord) -> None:
    if record.tv_license_eur and not re.fullmatch(r"\d+\.\d{2}", record.tv_license_eur):
        record.tv_license_eur = ""


def clean_acea_conguaglio_record(record: BillRecord) -> None:
    clean_acea_record(record)
    # In Acea conguaglio bills, detailed rows often mix historical recharge/reversal periods.
    # Keep reliable summary fields and blank misleading component-level rollups.
    detail_fields = [
        "energy_qty",
        "energy_unit_rate",
        "energy_imponibile_eur",
        "losses_qty",
        "losses_unit_rate",
        "losses_imponibile_eur",
        "dispbt_qty",
        "dispbt_unit_rate",
        "dispbt_imponibile_eur",
        "commercialization_qty",
        "commercialization_unit_rate",
        "commercialization_imponibile_eur",
        "capacity_market_qty",
        "capacity_market_unit_rate",
        "capacity_market_imponibile_eur",
        "dispatching_qty",
        "dispatching_unit_rate",
        "dispatching_imponibile_eur",
        "transport_energy_qty",
        "transport_energy_unit_rate",
        "transport_energy_imponibile_eur",
        "transport_fixed_qty",
        "transport_fixed_unit_rate",
        "transport_fixed_imponibile_eur",
        "transport_power_qty",
        "transport_power_unit_rate",
        "transport_power_imponibile_eur",
        "uc3_qty",
        "uc3_unit_rate",
        "uc3_imponibile_eur",
        "uc6_fixed_qty",
        "uc6_fixed_unit_rate",
        "uc6_fixed_imponibile_eur",
        "uc6_variable_qty",
        "uc6_variable_unit_rate",
        "uc6_variable_imponibile_eur",
        "arim_qty",
        "arim_unit_rate",
        "arim_imponibile_eur",
        "asos_qty",
        "asos_unit_rate",
        "asos_imponibile_eur",
        "excise_qty",
        "excise_unit_rate",
        "excise_imponibile_eur",
    ]
    for field in detail_fields:
        setattr(record, field, "")


def append_note(record: BillRecord, text: str) -> None:
    if not text:
        return
    if not record.notes:
        record.notes = text
    elif text not in record.notes:
        record.notes = f"{record.notes}; {text}"


def apply_generic_template(record: BillRecord, raw_text: str, lines: list[str]) -> None:
    for field, value in build_generic_regex_overrides(raw_text, lines).items():
        setattr(record, field, value)


def apply_octopus_template(record: BillRecord, raw_text: str, lines: list[str]) -> None:
    for field, value in build_octopus_regex_overrides(raw_text, lines).items():
        setattr(record, field, value)
    record.consumption_f1_kwh, record.consumption_f2_kwh, record.consumption_f3_kwh = extract_fascia_consumptions(raw_text)
    for field, value in build_detail_component_overrides(raw_text).items():
        setattr(record, field, value)
    clean_octopus_record(record)


def apply_acea_template(record: BillRecord, raw_text: str, lines: list[str]) -> None:
    for field, value in build_acea_regex_overrides(raw_text, lines).items():
        setattr(record, field, value)
    record.consumption_f1_kwh, record.consumption_f2_kwh, record.consumption_f3_kwh = extract_acea_fascia_consumptions(raw_text)
    for field, value in build_acea_detail_component_overrides(raw_text).items():
        setattr(record, field, value)
    clean_acea_record(record)


def apply_acea_conguaglio_template(record: BillRecord, raw_text: str, lines: list[str]) -> None:
    for field, value in build_acea_regex_overrides(raw_text, lines).items():
        setattr(record, field, value)
    record.consumption_f1_kwh, record.consumption_f2_kwh, record.consumption_f3_kwh = extract_acea_fascia_consumptions(raw_text)
    clean_acea_conguaglio_record(record)
    append_note(
        record,
        "Bolletta Acea con conguaglio: dettagli componente lasciati vuoti per evitare somme fuorvianti su periodi storici/ricalcoli",
    )


TEMPLATE_APPLIERS: dict[str, Callable[[BillRecord, str, list[str]], None]] = {
    "generic": apply_generic_template,
    "octopus": apply_octopus_template,
    "acea": apply_acea_template,
    "acea_standard": apply_acea_template,
    "acea_conguaglio": apply_acea_conguaglio_template,
}


def build_record(pdf_path: Path) -> BillRecord:
    raw_text = extract_pdf_text(pdf_path)
    lines = extract_lines(raw_text)
    record = BillRecord(source_file=str(pdf_path.name))
    record.supplier_template = infer_supplier_template(raw_text, lines)

    for field, labels in TEXT_LABELS.items():
        kind = "text"
        if field in {"invoice_date", "due_date"}:
            kind = "date"
        elif field == "pod_code":
            kind = "code"
        value = find_label_value(lines, labels, kind)
        setattr(record, field, value)

    for field, labels in MONEY_LABELS.items():
        setattr(record, field, find_label_value(lines, labels, "money"))

    record.supplier_name = record.supplier_name or infer_supplier(lines, pdf_path)
    record.billing_period_start, record.billing_period_end = find_period(lines, raw_text)
    record.consumption_kwh = find_consumption(lines, raw_text)
    record.committed_power_kw = find_committed_power(lines, raw_text)
    TEMPLATE_APPLIERS.get(record.supplier_template, apply_generic_template)(record, raw_text, lines)
    record.notes = build_notes(record)
    return record


def build_notes(record: BillRecord) -> str:
    missing = [key for key, value in asdict(record).items() if key not in {"source_file", "notes"} and not value]
    parts: list[str] = []
    if record.notes:
        parts.append(record.notes)
    if missing:
        parts.append("Campi mancanti: " + ", ".join(missing))
    return "; ".join(parts)


def discover_pdfs(path: Path) -> list[Path]:
    if path.is_file() and path.suffix.lower() == ".pdf":
        return [path]
    if path.is_dir():
        return sorted(file for file in path.rglob("*.pdf") if file.is_file())
    raise FileNotFoundError(f"Input non trovato: {path}")


def export_csv(records: list[BillRecord], output_path: Path) -> None:
    import re
    output_path.parent.mkdir(parents=True, exist_ok=True)
    # Prepara dati come per l'export Excel
    rows = [asdict(record) for record in records]
    if not rows:
        return
    columns = list(rows[0].keys())
    unit_map = {}
    for col in columns:
        if col.endswith('_unit_rate'):
            units = set()
            values = []
            for v in [str(row[col]) for row in rows]:
                m = re.match(r"([\d.,\-]+)\s*([\w€/\.]+)?", v.strip())
                if m:
                    values.append(m.group(1).replace(',', '.'))
                    if m.group(2):
                        units.add(m.group(2))
                    else:
                        units.add('')
                else:
                    values.append('')
                    units.add('')
            for i, row in enumerate(rows):
                row[col] = values[i]
            unit_map[col] = next(iter(units - {''}), '') if len(units - {''}) == 1 else ''
    columns_with_units = [col + (f" ({unit_map[col]})" if unit_map.get(col) else "") if col.endswith('_unit_rate') else col for col in columns]
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns_with_units)
        writer.writeheader()
        for row in rows:
            out_row = {}
            for col in columns:
                key = col + (f" ({unit_map[col]})" if unit_map.get(col) else "") if col.endswith('_unit_rate') else col
                out_row[key] = row[col]
            writer.writerow(out_row)


def export_xlsx(records: list[BillRecord], output_path: Path) -> None:
    import openpyxl
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment, NamedStyle
    from openpyxl.styles.numbers import FORMAT_NUMBER_00, FORMAT_DATE_DMYSLASH

    output_path.parent.mkdir(parents=True, exist_ok=True)
    frame = pd.DataFrame([asdict(record) for record in records], columns=OUTPUT_COLUMNS)

    # Gestione colonne unit_rate: estrai valore e unità
    unit_map = {}
    for col in frame.columns:
        if col.endswith('_unit_rate'):
            # Estrai unità e valore numerico
            units = set()
            values = []
            for v in frame[col].astype(str):
                m = re.match(r"([\d.,\-]+)\s*([\w€/\.]+)?", v.strip())
                if m:
                    values.append(m.group(1).replace(',', '.'))
                    if m.group(2):
                        units.add(m.group(2))
                    else:
                        units.add('')
                else:
                    values.append('')
                    units.add('')
            # Sostituisci i valori nella colonna solo col numero
            frame[col] = pd.to_numeric(values, errors="coerce")
            # Salva l'unità (se unica) per intestazione
            unit_map[col] = next(iter(units - {''}), '') if len(units - {''}) == 1 else ''

    # Numeric columns to float
    for column in NUMERIC_COLUMNS:
        if column in frame.columns:
            frame[column] = pd.to_numeric(frame[column].replace("", pd.NA), errors="coerce")
            frame[column] = frame[column].astype(float)

    # Date columns
    DATE_COLUMNS = [
        "invoice_date", "due_date", "billing_period_start", "billing_period_end"
    ]
    for column in DATE_COLUMNS:
        if column in frame.columns:
            frame[column] = pd.to_datetime(frame[column], errors="coerce")

    # Ordina per data inizio periodo fatturazione (crescente)
    if "billing_period_start" in frame.columns:
        frame["billing_period_start"] = pd.to_datetime(frame["billing_period_start"], errors="coerce")
        frame = frame.sort_values("billing_period_start", ascending=True)
    # Modifica intestazioni per unit_rate
    columns_with_units = [col + (f" ({unit_map[col]})" if unit_map.get(col) else "") if col.endswith('_unit_rate') else col for col in frame.columns]
    
    # Write to Excel con intestazioni modificate
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        frame.to_excel(writer, index=False, header=columns_with_units)

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Numeric formatting
    for col_idx, col_name in enumerate(frame.columns, 1):
        col_letter = get_column_letter(col_idx)
        if col_name in ["consumption_kwh", "consumption_f1_kwh", "consumption_f2_kwh", "consumption_f3_kwh"]:
            for cell in ws[col_letter]:
                if cell.row == 1:
                    continue
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="right")
        elif col_name in NUMERIC_COLUMNS:
            for cell in ws[col_letter]:
                if cell.row == 1:
                    continue
                cell.number_format = FORMAT_NUMBER_00
                cell.alignment = Alignment(horizontal="right")
        elif col_name in DATE_COLUMNS:
            for cell in ws[col_letter]:
                if cell.row == 1:
                    continue
                # Se la cella contiene anche orario, usa formato con ora
                if hasattr(cell.value, 'hour') and (cell.value.hour != 0 or cell.value.minute != 0 or cell.value.second != 0):
                    cell.number_format = 'DD/MM/YYYY HH:MM:SS'
                else:
                    cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal="center")
        else:
            for cell in ws[col_letter]:
                if cell.row == 1:
                    continue
                cell.alignment = Alignment(wrap_text=True, horizontal="left")

    # Auto-fit column widths
    for col_idx, col_name in enumerate(frame.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_length = max(
            [len(str(cell.value)) if cell.value is not None else 0 for cell in ws[col_letter]] + [len(col_name)]
        )
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)
    # Formattazione header: bold, sfondo grigio chiaro
    from openpyxl.styles import Font, PatternFill
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Imposta filtro sulla prima riga
    ws.auto_filter.ref = ws.dimensions

    # Imposta altezza minima per tutte le righe (eccetto header)
    min_height = 15
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = min_height
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Estrae dati chiave da bollette luce PDF e li salva in CSV o XLSX."
    )
    parser.add_argument("input_path", help="Percorso a un PDF o a una cartella contenente PDF.")
    parser.add_argument(
        "-o",
        "--output",
        default="output/bollette_estratte.csv",
        help="Percorso del file di output. Estensioni supportate: .csv, .xlsx",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    input_path = Path(args.input_path).expanduser().resolve()
    output_path = Path(args.output).expanduser()
    if not output_path.is_absolute():
        output_path = Path.cwd() / output_path
    output_path = output_path.resolve()

    pdfs = discover_pdfs(input_path)
    if not pdfs:
        raise SystemExit("Nessun PDF trovato nel percorso indicato.")

    records = [build_record(pdf_path) for pdf_path in pdfs]

    suffix = output_path.suffix.lower()
    if suffix == ".csv":
        export_csv(records, output_path)
    elif suffix == ".xlsx":
        export_xlsx(records, output_path)
    else:
        raise SystemExit("Formato output non supportato. Usa .csv oppure .xlsx")

    print(f"Creato: {output_path}")
    print(f"PDF elaborati: {len(records)}")


if __name__ == "__main__":
    main()
