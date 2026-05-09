from __future__ import annotations

import re
from dataclasses import asdict
from datetime import date, datetime
from pathlib import Path
from zipfile import ZIP_DEFLATED, ZipFile
from xml.etree import ElementTree as ET

import pandas as pd

from .models import ELECTRICITY_SERVICE_TYPE, GAS_SERVICE_TYPE
from .electricity.models import BillRecord, NUMERIC_COLUMNS, OUTPUT_COLUMNS
from .gas.models import GasBillRecord, GAS_NUMERIC_COLUMNS, GAS_OUTPUT_COLUMNS
from .output_config import OutputColumn, default_output_columns


DATE_COLUMNS = ["invoice_date", "due_date", "billing_period_start", "billing_period_end"]
SUPPORTED_EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
Record = BillRecord | GasBillRecord

# (sheet_name, records, columns, service_type)
SheetSpec = tuple[str, list[Record], list[OutputColumn] | None, str]


def export_xlsx(sheets: list[SheetSpec], output_path: Path) -> None:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.styles.numbers import FORMAT_NUMBER_00
    from openpyxl.utils import get_column_letter

    output_path.parent.mkdir(parents=True, exist_ok=True)
    suffix = output_path.suffix.lower()
    if suffix == ".xls":
        raise ValueError("Formato .xls non supportato. Usa .xlsx o .xlsm.")
    if suffix not in SUPPORTED_EXCEL_SUFFIXES:
        raise ValueError("Formato output non supportato. Usa .xlsx o .xlsm.")
    if suffix == ".xlsm" and output_path.exists():
        _export_existing_xlsm(sheets, output_path)
        return

    existing_sheets: dict[str, pd.DataFrame] = {}
    if output_path.exists():
        try:
            with pd.ExcelFile(output_path) as excel_file:
                existing_sheets = pd.read_excel(excel_file, sheet_name=None)
        except Exception:
            existing_sheets = {}

    # Build one PreparedSheet per sheet before writing
    prepared: list[_PreparedSheet] = [
        _prepare_sheet(
            sheet_name,
            records,
            columns,
            service_type,
            existing_sheets,
        )
        for sheet_name, records, columns, service_type in sheets
    ]

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path, keep_vba=suffix == ".xlsm")
    else:
        wb = Workbook()
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

    for ps in prepared:
        if ps.sheet_name in wb.sheetnames:
            ws = wb[ps.sheet_name]
            _clear_sheet(ws)
        else:
            ws = wb.create_sheet(ps.sheet_name)
        _write_prepared_sheet(ws, ps)

    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    header_font = Font(bold=True)

    for ps in prepared:
        ws = wb[ps.sheet_name]

        for col_idx, col_name in enumerate(ps.frame.columns, 1):
            col_letter = get_column_letter(col_idx)
            for cell in ws[col_letter]:
                if cell.row == 1:
                    continue
                if col_name in (
                    "consumption_kwh",
                    "consumption_f1_kwh",
                    "consumption_f2_kwh",
                    "consumption_f3_kwh",
                    "reactive_energy_kvarh",
                    "reactive_energy_f1_kvarh",
                    "reactive_energy_f2_kvarh",
                    "reactive_energy_f3_kvarh",
                    "consumption_smc",
                    "estimated_consumption_smc",
                    "annual_consumption_smc",
                ):
                    cell.number_format = "0"
                    cell.alignment = Alignment(horizontal="right")
                elif col_name in ps.numeric_columns:
                    cell.number_format = FORMAT_NUMBER_00
                    cell.alignment = Alignment(horizontal="right")
                elif col_name in DATE_COLUMNS:
                    has_time = hasattr(cell.value, "hour") and (
                        cell.value.hour or cell.value.minute or cell.value.second
                    )
                    cell.number_format = "DD/MM/YYYY HH:MM:SS" if has_time else "DD/MM/YYYY"
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(wrap_text=True, horizontal="left")

        for col_idx, col_name in enumerate(ps.frame.columns, 1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(c.value)) if c.value is not None else 0 for c in ws[col_letter]),
                default=0,
            )
            ws.column_dimensions[col_letter].width = min(max(max_len, len(col_name)) + 2, 40)

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin
                if cell.row == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 15

        ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    wb.close()


def _clear_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None


def _write_prepared_sheet(
    ws,
    prepared: _PreparedSheet,
    *,
    write_headers: bool = True,
) -> None:
    if write_headers:
        for col_idx, header in enumerate(prepared.headers, 1):
            ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(prepared.frame.itertuples(index=False, name=None), 2):
        for col_idx, value in enumerate(row, 1):
            if pd.isna(value):
                value = None
            ws.cell(row=row_idx, column=col_idx, value=value)


class _PreparedSheet:
    __slots__ = ("sheet_name", "frame", "headers", "numeric_columns")

    def __init__(
        self,
        sheet_name: str,
        frame: pd.DataFrame,
        headers: list[str],
        numeric_columns: list[str],
    ) -> None:
        self.sheet_name = sheet_name
        self.frame = frame
        self.headers = headers
        self.numeric_columns = numeric_columns


def _prepare_sheet(
    sheet_name: str,
    records: list[Record],
    columns: list[OutputColumn] | None,
    service_type: str,
    existing_sheets: dict[str, pd.DataFrame],
    forced_column_names: list[str] | None = None,
    forced_headers: list[str] | None = None,
) -> _PreparedSheet:
    output_columns = columns or default_output_columns(service_type)
    column_names = forced_column_names or [col.source for col in output_columns]
    model_columns = _model_columns(service_type)
    numeric_columns = _numeric_columns(service_type)

    frame = pd.DataFrame([asdict(r) for r in records], columns=model_columns)
    frame = frame[column_names]

    unit_map: dict[str, str] = {}
    for col in frame.columns:
        if not col.endswith("_unit_rate"):
            continue
        units: set[str] = set()
        values: list[str] = []
        for v in frame[col].astype(str):
            m = re.match(r"([\d.,\-]+)\s*([\w€/\.]+)?", v.strip())
            if m:
                values.append(m.group(1).replace(",", "."))
                units.add(m.group(2) or "")
            else:
                values.append("")
                units.add("")
        frame[col] = pd.to_numeric(values, errors="coerce")
        unit_map[col] = next(iter(units - {""}), "") if len(units - {""}) == 1 else ""

    for col in numeric_columns:
        if col in frame.columns:
            frame[col] = pd.to_numeric(frame[col].replace("", pd.NA), errors="coerce").astype(float)

    for col in DATE_COLUMNS:
        if col in frame.columns:
            frame[col] = pd.to_datetime(frame[col], errors="coerce")

    headers = forced_headers or [
        _col_header(col, unit_map)
        for col in output_columns
        if col.source in column_names
    ]
    if len(headers) != len(column_names):
        headers = column_names

    if sheet_name in existing_sheets and "source_file" in frame.columns:
        existing = existing_sheets[sheet_name].copy()
        header_to_source = dict(zip(headers, column_names, strict=False))
        existing = existing.rename(columns=header_to_source)
        if "source_file" in existing.columns:
            existing = existing.reindex(columns=column_names)
            frame = pd.concat([existing, frame], ignore_index=True)
            frame = frame.drop_duplicates(subset=["source_file"], keep="last")
            frame = frame.reindex(columns=column_names)

    if forced_column_names is None and "billing_period_start" in frame.columns:
        frame = frame.sort_values("billing_period_start", ascending=True)

    return _PreparedSheet(sheet_name, frame, headers, numeric_columns)


def _export_existing_xlsm(sheets: list[SheetSpec], output_path: Path) -> None:
    with ZipFile(output_path, "r") as archive:
        sheet_paths = _workbook_sheet_paths(archive)
        shared_strings = _read_shared_strings(archive)
        existing_sheets: dict[str, pd.DataFrame] = {}
        forced_columns: dict[str, list[str]] = {}
        forced_headers: dict[str, list[str]] = {}
        sheet_xml_by_name: dict[str, bytes] = {}

        for sheet_name, _records, columns, service_type in sheets:
            sheet_path = sheet_paths.get(sheet_name)
            if not sheet_path:
                raise ValueError(f"Il file .xlsm non contiene il foglio '{sheet_name}'.")
            sheet_xml = archive.read(sheet_path)
            sheet_xml_by_name[sheet_name] = sheet_xml
            column_names, headers = _worksheet_schema_from_xml(
                sheet_xml,
                shared_strings,
                columns,
                service_type,
            )
            if not column_names:
                raise ValueError(f"Impossibile leggere le intestazioni del foglio '{sheet_name}'.")
            forced_columns[sheet_name] = column_names
            forced_headers[sheet_name] = headers
            existing_sheets[sheet_name] = _worksheet_to_frame_from_xml(sheet_xml, shared_strings, column_names)

        prepared = [
            _prepare_sheet(
                sheet_name,
                records,
                columns,
                service_type,
                existing_sheets,
                forced_columns[sheet_name],
                forced_headers[sheet_name],
            )
            for sheet_name, records, columns, service_type in sheets
        ]
        prepared_by_path = {
            sheet_paths[ps.sheet_name]: _write_sheet_xml(sheet_xml_by_name[ps.sheet_name], ps)
            for ps in prepared
        }

        tmp_path = output_path.with_suffix(".tmp.xlsm")
        with ZipFile(tmp_path, "w", ZIP_DEFLATED) as target:
            for item in archive.infolist():
                data = prepared_by_path.get(item.filename)
                target.writestr(item, data if data is not None else archive.read(item.filename))
        tmp_path.replace(output_path)


def _workbook_sheet_paths(archive: ZipFile) -> dict[str, str]:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_targets = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall(f"{{{rel_ns}}}Relationship")
        if rel.attrib.get("Type", "").endswith("/worksheet")
    }
    paths: dict[str, str] = {}
    for sheet in workbook.findall(f"{{{main_ns}}}sheets/{{{main_ns}}}sheet"):
        rel_id = sheet.attrib.get(f"{{{office_rel_ns}}}id")
        target = rel_targets.get(rel_id or "")
        if target:
            paths[sheet.attrib["name"]] = _normalize_xl_target(target)
    return paths


def _normalize_xl_target(target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    if target.startswith("xl/"):
        return target
    return f"xl/{target}"


def _read_shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    values: list[str] = []
    for item in root.findall(f"{{{main_ns}}}si"):
        texts = [node.text or "" for node in item.findall(f".//{{{main_ns}}}t")]
        values.append("".join(texts))
    return values


def _worksheet_schema_from_xml(
    sheet_xml: bytes,
    shared_strings: list[str],
    columns: list[OutputColumn] | None,
    service_type: str,
) -> tuple[list[str], list[str]]:
    model_columns = set(_model_columns(service_type))
    configured_columns = columns or default_output_columns(service_type)
    header_to_source = {col.source: col.source for col in configured_columns}
    header_to_source.update({col.title: col.source for col in configured_columns if col.title})
    header_to_source.update({col: col for col in model_columns})
    first_row = _worksheet_row_values(sheet_xml, shared_strings, row_number=1)

    column_names: list[str] = []
    headers: list[str] = []
    for col_idx in range(1, max(first_row.keys(), default=0) + 1):
        header = str(first_row.get(col_idx) or "").strip()
        if not header:
            break
        source = header_to_source.get(header)
        if not source or source not in model_columns:
            break
        column_names.append(source)
        headers.append(header)
    return column_names, headers


def _worksheet_to_frame_from_xml(
    sheet_xml: bytes,
    shared_strings: list[str],
    column_names: list[str],
) -> pd.DataFrame:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    root = ET.fromstring(sheet_xml)
    rows: list[list[object]] = []
    for row in root.findall(f".//{{{main_ns}}}sheetData/{{{main_ns}}}row"):
        row_number = int(row.attrib.get("r", "0") or 0)
        if row_number < 2:
            continue
        values = [None] * len(column_names)
        for cell in row.findall(f"{{{main_ns}}}c"):
            col_idx = _cell_col_index(cell.attrib.get("r", ""))
            if 1 <= col_idx <= len(column_names):
                values[col_idx - 1] = _cell_value(cell, shared_strings)
        if any(value is not None for value in values):
            rows.append(values)
    return pd.DataFrame(rows, columns=column_names)


def _worksheet_row_values(sheet_xml: bytes, shared_strings: list[str], row_number: int) -> dict[int, object]:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    root = ET.fromstring(sheet_xml)
    values: dict[int, object] = {}
    for row in root.findall(f".//{{{main_ns}}}sheetData/{{{main_ns}}}row"):
        if int(row.attrib.get("r", "0") or 0) != row_number:
            continue
        for cell in row.findall(f"{{{main_ns}}}c"):
            col_idx = _cell_col_index(cell.attrib.get("r", ""))
            if col_idx:
                values[col_idx] = _cell_value(cell, shared_strings)
        break
    return values


def _cell_value(cell: ET.Element, shared_strings: list[str]) -> object:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    cell_type = cell.attrib.get("t")
    if cell_type == "s":
        value = cell.find(f"{{{main_ns}}}v")
        if value is None or value.text is None:
            return None
        index = int(value.text)
        return shared_strings[index] if 0 <= index < len(shared_strings) else None
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.findall(f".//{{{main_ns}}}t")) or None
    value = cell.find(f"{{{main_ns}}}v")
    if value is None or value.text is None:
        return None
    if cell_type == "b":
        return value.text == "1"
    if cell_type in {"str", "e"}:
        return value.text
    try:
        number = float(value.text)
    except ValueError:
        return value.text
    return int(number) if number.is_integer() else number


def _write_sheet_xml(sheet_xml: bytes, prepared: _PreparedSheet) -> bytes:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ET.register_namespace("", main_ns)
    ET.register_namespace("r", rel_ns)
    root = ET.fromstring(sheet_xml)
    sheet_data = root.find(f"{{{main_ns}}}sheetData")
    if sheet_data is None:
        sheet_data = ET.SubElement(root, f"{{{main_ns}}}sheetData")

    style_by_col = _column_styles_from_sheet_xml(sheet_xml)
    for row in list(sheet_data):
        if int(row.attrib.get("r", "0") or 0) >= 2:
            sheet_data.remove(row)

    for row_idx, values in enumerate(prepared.frame.itertuples(index=False, name=None), 2):
        row_el = ET.SubElement(
            sheet_data,
            f"{{{main_ns}}}row",
            {"r": str(row_idx), "ht": "15", "customHeight": "1"},
        )
        for col_idx, value in enumerate(values, 1):
            if pd.isna(value):
                continue
            cell_ref = f"{_col_letter(col_idx)}{row_idx}"
            attrs = {"r": cell_ref}
            if style := style_by_col.get(col_idx):
                attrs["s"] = style
            cell_el = ET.SubElement(row_el, f"{{{main_ns}}}c", attrs)
            _set_cell_xml_value(cell_el, value)

    max_row = max(len(prepared.frame) + 1, 1)
    max_col = max(len(prepared.frame.columns), 1)
    dimension = root.find(f"{{{main_ns}}}dimension")
    if dimension is not None:
        dimension.set("ref", f"A1:{_col_letter(max_col)}{max_row}")

    auto_filter = root.find(f"{{{main_ns}}}autoFilter")
    if auto_filter is not None:
        auto_filter.set("ref", f"A1:{_col_letter(max_col)}{max_row}")

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _column_styles_from_sheet_xml(sheet_xml: bytes) -> dict[int, str]:
    styles: dict[int, str] = {}
    for coord, style in re.findall(r'<c\b(?=[^>]*\br="([^"]+)")(?=[^>]*\bs="(\d+)")[^>]*>', sheet_xml.decode("utf-8")):
        row_match = re.search(r"\d+", coord)
        if not row_match:
            continue
        row_num = int(row_match.group(0))
        col_idx = _cell_col_index(coord)
        if row_num == 2:
            styles[col_idx] = style
        elif row_num == 1 and col_idx not in styles:
            styles[col_idx] = style
    return styles


def _set_cell_xml_value(cell_el: ET.Element, value: object) -> None:
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime):
        ET.SubElement(cell_el, f"{{{main_ns}}}v").text = _format_number(_excel_serial(value))
    elif isinstance(value, date):
        ET.SubElement(cell_el, f"{{{main_ns}}}v").text = _format_number(_excel_serial(datetime.combine(value, datetime.min.time())))
    elif isinstance(value, bool):
        cell_el.set("t", "b")
        ET.SubElement(cell_el, f"{{{main_ns}}}v").text = "1" if value else "0"
    elif isinstance(value, (int, float)) and not isinstance(value, bool):
        ET.SubElement(cell_el, f"{{{main_ns}}}v").text = _format_number(float(value))
    else:
        cell_el.set("t", "inlineStr")
        inline = ET.SubElement(cell_el, f"{{{main_ns}}}is")
        ET.SubElement(inline, f"{{{main_ns}}}t").text = str(value)


def _excel_serial(value: datetime) -> float:
    base = datetime(1899, 12, 30)
    delta = value.replace(tzinfo=None) - base
    return delta.days + (delta.seconds + delta.microseconds / 1_000_000) / 86400


def _format_number(value: float) -> str:
    return str(int(value)) if float(value).is_integer() else repr(float(value))


def _cell_col_index(cell_ref: str) -> int:
    letters = re.match(r"([A-Z]+)", cell_ref.upper())
    if not letters:
        return 0
    result = 0
    for char in letters.group(1):
        result = result * 26 + ord(char) - ord("A") + 1
    return result


def _col_letter(index: int) -> str:
    letters = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        letters = chr(ord("A") + remainder) + letters
    return letters


def _col_header(output_col: OutputColumn, unit_map: dict[str, str]) -> str:
    col = output_col.source
    title = output_col.title or col
    if output_col.title:
        return title
    return title + (f" ({unit_map[col]})" if unit_map.get(col) else "") if col.endswith("_unit_rate") else title


def _model_columns(service_type: str) -> list[str]:
    if service_type == ELECTRICITY_SERVICE_TYPE:
        return OUTPUT_COLUMNS
    if service_type == GAS_SERVICE_TYPE:
        return GAS_OUTPUT_COLUMNS
    raise ValueError(f"Tipo servizio non supportato: {service_type}")


def _numeric_columns(service_type: str) -> list[str]:
    if service_type == ELECTRICITY_SERVICE_TYPE:
        return NUMERIC_COLUMNS
    if service_type == GAS_SERVICE_TYPE:
        return GAS_NUMERIC_COLUMNS
    raise ValueError(f"Tipo servizio non supportato: {service_type}")
